from django.contrib import admin
from django.contrib.auth.hashers import make_password
from django.template.response import TemplateResponse
from django.urls import reverse
from user.models.sms import OTP, UserSMS
from user.models.user import User
from user.models.student import Student
from university.models.application import StudentApplication, Application
from django.utils.translation import gettext_lazy as _
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, HttpResponseForbidden
from django.urls import path
from django.utils.html import format_html
from university.services.generate_contract.generate_contract import generate_student_contract, generate_student_certificate
from django.shortcuts import redirect
from django.contrib import messages
from django.utils.safestring import mark_safe
from django.db import models
from django.http import FileResponse, Http404
from university.models.student_dtm import StudentDtmResult
import os
class SafeBaseModelAdmin(admin.ModelAdmin):
    def get_queryset(self, request):
        return self.model.all_objects.all()
    list_display = ['__str__', 'is_active', 'created_datetime', 'modified_datetime']
    list_filter = ['is_active']
    search_fields = ['id']
    actions = ['make_active', 'make_inactive']

    def make_active(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f"{updated} items marked as active.")
    make_active.short_description = "Mark selected as active"

    def make_inactive(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f"{updated} items marked as inactive.")
    make_inactive.short_description = "Mark selected as inactive"

def export_users_as_excel(modeladmin, request, queryset):
    fields = [
        ('ID', 'id'),
        ('Phone Number', 'phone_number'),
        ('Is Verified', 'is_verified'),
        ('Is Active', 'is_active'),
        ('Created', 'created_datetime'),
        ('Modified', 'modified_datetime'),
        ('UUID', 'uuid'),
        # Add more user fields here as needed
    ]
    # Add region if User has a region field
    if hasattr(User, 'region'):
        fields.insert(3, ('Region', lambda obj: obj.region.name if obj.region else ''))

    # Add full_name if present or make a property:
    if hasattr(User, 'full_name'):
        fields.insert(2, ('Full Name', 'full_name'))
    else:
        fields.insert(2, ('Full Name', lambda obj: f"{getattr(obj, 'first_name', '')} {getattr(obj, 'last_name', '')}".strip()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    for col_num, (header, _) in enumerate(fields, 1):
        ws.cell(row=1, column=col_num, value=header)
    for row_num, user in enumerate(queryset, 2):
        for col_num, (_, attr) in enumerate(fields, 1):
            value = attr(user) if callable(attr) else getattr(user, attr, "")
            ws.cell(row=row_num, column=col_num, value=str(value))
    for i in range(1, len(fields) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=users_export.xlsx'
    wb.save(response)
    return response
export_users_as_excel.short_description = "Export selected users to Excel"

def export_students_as_excel(modeladmin, request, queryset):
    fields = [
        ('ID', 'id'),
        ('First Name', 'first_name'),
        ('Last Name', 'last_name'),
        ('Father Name', 'father_name'),
        ('Additional Phone Number', 'additional_phone_number'),
        ('Birth Date', 'birth_date'),
        ('Birth Place', 'birth_place'),
        ('Citizenship', 'citizenship'),
        ('PINFL', 'pinfl'),
        ('Passport Number', 'passport_number'),
        ('Gender', 'gender'),
        ('Qualification', 'qualification'),
        ('Name Qualification', 'name_qualification'),
        ('Field of Study', lambda obj: str(
            obj.applications.first().study_type) if obj.applications.exists() and obj.applications.first().study_type else ''),
        ('Form of Education', lambda obj: str(
            obj.applications.first().faculty) if obj.applications.exists() and obj.applications.first().faculty else ''),
        ('Programs of study', lambda obj: str(obj.applications.first().program) if obj.applications.exists() and obj.applications.first().program else ''),
        ('Language of study', lambda obj: str(obj.applications.first().get_lang_display()) if obj.applications.exists() and obj.applications.first().lang else ''),
        ('Is Registered', 'is_registred'),
        ('Is Attended Exam', 'is_attended_exam'),
        ('Is Passed Exam', 'is_passed_exam'),
        ('Last Login', 'last_login'),
        ('Is Blocked', 'is_blocked'),
        ('Is Active', 'is_active'),
        ('Created', 'created_datetime'),
        ('Modified', 'modified_datetime'),
        ('UUID', 'uuid'),
    ]
    # Add region if Student has a region field
    if hasattr(Student, 'region'):
        fields.insert(3, ('Region', lambda obj: obj.region.name if obj.region else ''))
    # Add user fields if needed
    fields.insert(4, ('User Phone', lambda obj: obj.user.phone_number if obj.user else ''))
    fields.insert(5, ('User Full Name', lambda obj: obj.user.full_name if hasattr(obj.user, "full_name") else (
        f"{getattr(obj.user, 'first_name', '')} {getattr(obj.user, 'last_name', '')}".strip() if obj.user else ''
    )))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    for col_num, (header, _) in enumerate(fields, 1):
        ws.cell(row=1, column=col_num, value=header)
    for row_num, obj in enumerate(queryset, 2):
        for col_num, (_, attr) in enumerate(fields, 1):
            try:
                value = attr(obj) if callable(attr) else getattr(obj, attr, "")
            except Exception:
                value = ''
            ws.cell(row=row_num, column=col_num, value=str(value))
    for i in range(1, len(fields) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=students_export.xlsx'
    wb.save(response)
    return response
export_students_as_excel.short_description = "Export selected students to Excel"

# --------- Inlines ----------
class StudentApplicationInline(admin.TabularInline):
    model = StudentApplication
    extra = 0
    show_change_link = True

# --------- Admins ---------
@admin.register(User)
class UserAdmin(SafeBaseModelAdmin):
    list_display = [
        'id', 'phone_number', 'source', 'registered_by_bot', 'is_verified', 'is_active', 'created_datetime'
    ]
    # Add region or full_name if present to list_display as well
    if hasattr(User, 'region'):
        list_display.insert(3, 'region')
    if hasattr(User, 'full_name'):
        list_display.insert(2, 'full_name')

    search_fields = ['phone_number', 'id', 'source']
    list_filter = ['is_verified', 'is_active', 'source', 'registered_by_bot']
    if hasattr(User, 'region'):
        list_filter.append('region')
    readonly_fields = ['created_datetime', 'modified_datetime', 'uuid', 'source', 'registered_by_bot']
    actions = [export_users_as_excel]

    def save_model(self, request, obj, form, change):
        # Only hash if the password is not already hashed
        if obj.password and not obj.password.startswith('pbkdf2_'):
            obj.password = make_password(obj.password)
        super().save_model(request, obj, form, change)

@admin.register(Student)
class StudentAdmin(SafeBaseModelAdmin):
    list_display = [
        'id',
        'first_name',
        'last_name',
        'user',
        *(['region'] if hasattr(Student, 'region') else []),
        'is_registred',
        'is_active',
        'created_datetime',
        'contract_certificate_buttons',  # single column with both buttons
    ]
    search_fields = ['first_name', 'pinfl', 'passport_number', 'last_name', 'user__phone_number']
    list_filter = [
        'is_registred',
        'is_active',
        *(['region'] if hasattr(Student, 'region') else [])
    ]

    actions = [export_students_as_excel]
    inlines = [StudentApplicationInline]

    def get_readonly_fields(self, request, obj=None):
        if obj:  # edit mode
            return [
                'created_datetime',
                'modified_datetime',
                'uuid',
                'download_diploma',
                'download_photo',
                'download_dtm',
            ]
        else:  # create mode
            return [
                'created_datetime',
                'modified_datetime',
                'uuid',
                'download_diploma',
                'download_photo',
                'download_dtm',
            ]

    def download_diploma(self, obj):
        if obj.diploma:
            url = reverse('admin:download-student-file', args=['diploma', obj.id])
            return format_html(f'<a class="button" href="{url}">🎓 Diplomani yuklab olish</a>')
        return "Diploma mavjud emas"

    download_diploma.short_description = "Diploma"

    def download_photo(self, obj):
        if obj.photo:
            url = reverse('admin:download-student-file', args=['photo', obj.id])
            return format_html(f'<a class="button" href="{url}">🖼 Rasmni yuklab olish</a>')
        return "Rasm mavjud emas"
    def download_dtm(self, obj):
        student_dtm = StudentDtmResult.objects.filter(student_id=obj.id).first()
        if student_dtm and student_dtm.dtm_file:
            url = reverse('admin:download-student-dtm', args=[obj.id])
            return format_html(f'<a class="button" href="{url}">🧾 DTM blankni yuklab olish</a>')
        return "DTM blank mavjud emas"


    download_photo.short_description = "Rasm"
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'gender' in form.base_fields:
            form.base_fields.pop('gender')
        return form

    def gender_display(self, obj):
        if not obj.gender:
            return "-"
        gender_map = {'1': 'Male', '2': 'Female', 1: 'Male', 2: 'Female'}
        return gender_map.get(obj.gender, '-')
    gender_display.short_description = "Gender"

    def contract_certificate_buttons(self, obj):
        buttons_html = ""
        args = []

        # Kontrakt va sertifikat faqat imtihon topshirgan va o‘tganlar uchun
        if obj.is_attended_exam and obj.is_passed_exam:
            contract_url = f"generate_contract/{obj.id}/"
            certificate_url = f"generate_certificate/{obj.id}/"
            buttons_html += (
                '<a class="button" href="{}" target="_blank">📄 Yuklab olish</a>&nbsp;'
                '<a class="button" href="{}" target="_blank">📜 Sertifikat</a>'
            )
            args += [contract_url, certificate_url]

        # DTM fayl bo‘lsa, har doim chiqsin
        student_dtm = StudentDtmResult.objects.filter(student_id=obj.id).first()
        if student_dtm and student_dtm.dtm_file:
            dtm_url = f"download_dtm/{obj.id}/"
            buttons_html += '&nbsp;<a class="button" href="{}" target="_blank">🧾 DTM blank</a>'
            args.append(dtm_url)

        return format_html(buttons_html, *args) if buttons_html else "-"

    contract_certificate_buttons.short_description = "Actions"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
    'download_dtm/<int:student_id>/',
    self.admin_site.admin_view(self.download_dtm_file_view),
    name='download-student-dtm',
),
            path(
                'download/<str:file_type>/<int:student_id>/',
                self.admin_site.admin_view(self.download_file_view),
                name='download-student-file',
            ),
            path(
                'generate_contract/<int:student_id>/',
                self.admin_site.admin_view(self.generate_contract_pdf_view),
                name='generate-student-contract',
            ),
            path(
                'generate_certificate/<int:student_id>/',
                self.admin_site.admin_view(self.generate_certificate_docx_view),
                name='generate-student-certificate',
            ),
        ]
        return custom_urls + urls

    def download_file_view(self, request, file_type, student_id):
        from django.http import FileResponse, Http404
        import os

        student = Student.objects.filter(id=student_id).first()
        if not student:
            raise Http404("Talaba topilmadi")

        if file_type == 'diploma':
            file_field = student.diploma
        elif file_type == 'photo':
            file_field = student.photo
        else:
            raise Http404("Noto‘g‘ri fayl turi")

        if not file_field:
            raise Http404("Fayl mavjud emas")

        file_path = file_field.path
        if not os.path.exists(file_path):
            raise Http404("Fayl mavjud emas (fizik)")

        response = FileResponse(open(file_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response
    def generate_contract_pdf_view(self, request, student_id):
        language = request.GET.get("language")
        if language:
            try:
                return generate_student_contract(student_id, language=language)
            except Exception as e:
                self.message_user(request, f"Xatolik yuz berdi: {e}", level=messages.ERROR)
                return redirect(request.META.get('HTTP_REFERER', '/admin/'))

        # Show language selection message
        url_base = request.path
        msg = (
            f"Tanishuv: Shartnoma tili tanlanmagan. "
            f"Tanlang: <a href='{url_base}?language=uz'>📄 UZ</a> | "
            f"<a href='{url_base}?language=ru'>📄 RU</a>"
        )
        self.message_user(request, mark_safe(msg), level=messages.INFO)
        return redirect(request.META.get('HTTP_REFERER', '/admin/'))

    def generate_certificate_docx_view(self, request, student_id):
        language = request.GET.get("language")
        if language:
            try:
                response = generate_student_certificate(student_id, language=language)
                return response  # this must return the HttpResponse with the file
            except Exception as e:
                self.message_user(request, f"Xatolik yuz berdi: {e}", level=messages.ERROR)
                return redirect(request.META.get('HTTP_REFERER', '/admin/'))

        url_base = request.path
        msg = (
            f"Tanishuv: Sertifikat tili tanlanmagan. "
            f"Tanlang: <a href='{url_base}?language=uz'>📜 UZ</a> | "
            f"<a href='{url_base}?language=ru'>📜 RU</a>"
        )
        self.message_user(request, mark_safe(msg), level=messages.INFO)
        return redirect(request.META.get('HTTP_REFERER', '/admin/'))
    def download_dtm_file_view(self, request, student_id):
        

        dtm_result = StudentDtmResult.objects.filter(student_id=student_id).first()
        if not dtm_result or not dtm_result.dtm_file:
            raise Http404("DTM blank fayli topilmadi")

        file_path = dtm_result.dtm_file.path
        if not os.path.exists(file_path):
            raise Http404("Fayl fizik jihatdan mavjud emas")

        response = FileResponse(open(file_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response

    

@admin.register(OTP)
class OTPAdmin(SafeBaseModelAdmin):
    list_display = ['id', 'phone_number', 'code', 'is_verified', 'expires_at', 'is_active']
    search_fields = ['phone_number', 'code']
    list_filter = ['is_verified', 'is_active']

@admin.register(UserSMS)
class UserSMSAdmin(SafeBaseModelAdmin):
    list_display = ['id', 'user', 'category', 'is_sent', 'is_active', 'created_datetime']
    search_fields = ['user__phone_number', 'text']
    list_filter = ['is_sent', 'category', 'is_active']




class CustomHTMLAdmin(admin.ModelAdmin):
    allowed_usernames = ["Operator", "nodirbek"]
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('', self.admin_site.admin_view(self.custom_view), name='admin-custom-html'),
        ]
        return custom_urls + urls

    def custom_view(self, request):

        if request.user.username not in self.allowed_usernames:
            return HttpResponseForbidden("Sizda bu sahifaga ruxsat yo‘q.")
        context = dict(
            self.admin_site.each_context(request),
            title='Custom HTML Sahifa',

        )
        application = Application.objects.prefetch_related('programs').first()
        application_faculty = Application.objects.prefetch_related('faculties').first()
        application_study_types = Application.objects.prefetch_related('study_types').first()
        context_value={
            "programs": application.programs.all(),
            "faculty": application_faculty.faculties.all(),
            "study_types": application_study_types.study_types.all(),
        }
        return TemplateResponse(request, "user/student_form.html",context_value)
class DummyModel(models.Model):
    class Meta:
        verbose_name_plural = "📄 Admin Student register"
        managed = False

admin.site.register(DummyModel, CustomHTMLAdmin)

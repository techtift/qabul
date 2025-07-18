from django.urls import path, reverse
from modeltranslation.admin import TranslationAdmin
from django.contrib import admin
from university.models.application import Application, StudentApplication
from university.models.duration import Duration
from university.models.exam import Exam
from university.models.faculty import Faculty
from university.models.program import Program
from university.models.study_type import StudyType
from university.models.subject import Subject
from university.models.study_type_many_to_many import StudyTypeFaculty
from university.models.question import Question
from university.models.exam_type import ExamType
from django.db import models
from django.utils.html import format_html
from university.translation import *
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from rangefilter.filters import DateRangeFilter
from django.http import FileResponse, Http404
import os
from django.core.exceptions import PermissionDenied


class SafeBaseModelAdmin(admin.ModelAdmin):
    list_display = ['__str__', 'is_active', 'created_datetime', 'modified_datetime', 'uuid']
    list_filter = ['is_active', 'created_datetime']
    search_fields = ['id', 'uuid']
    readonly_fields = ['uuid', 'created_datetime', 'modified_datetime']
    date_hierarchy = 'created_datetime'
    actions = ['make_active', 'make_inactive']

    def make_active(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f"{updated} items marked as active.")

    make_active.short_description = "Mark selected as active"

    def make_inactive(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f"{updated} items marked as inactive.")

    make_inactive.short_description = "Mark selected as inactive"


@admin.action(description="Export registered students")
def export_registered_students(modeladmin, request, queryset):
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registered Students"

    # ✅ Read date filters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    start_date = None
    end_date = None
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        if end_date_str:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
    except ValueError:
        pass

    # ✅ Show date range info at the top
    date_range_text = "Sana oraliği: "
    if start_date:
        date_range_text += f"{start_date.strftime('%Y-%m-%d')} dan "
    if end_date:
        date_range_text += f"{end_date.strftime('%Y-%m-%d')} gacha"
    if not start_date and not end_date:
        date_range_text += "Tanlanmagan"

    ws.merge_cells('A1:H1')
    ws['A1'] = date_range_text

    # ✅ Table headers
    headers = [
        "№", "F.I.O", "Passport", "Telefon", "Tug‘ilgan sana",
        "Ro'yxatdan o'tgan sana", "Ta'lim yo'nalishi", "Shakli", "Imtihon topshirdi"
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=2, column=col_num, value=header)

    # ✅ Collect StudentApplications filtered by selected faculties and date range
    applications = StudentApplication.objects.filter(
        faculty__in=queryset, is_active=True
    ).select_related("faculty", "study_type", "student").order_by('created_datetime')

    if start_date:
        applications = applications.filter(created_datetime__gte=start_date)
    if end_date:
        applications = applications.filter(created_datetime__lte=end_date)

    for idx, app in enumerate(applications, start=1):
        student = app.student
        row = [
            idx,
            student.first_name + " " + student.last_name + " " + (student.father_name or ""),
            student.passport_number if student else "",
            student.user.phone_number if student else "",
            student.birth_date.strftime('%Y-%m-%d') if student and student.birth_date else "",
            app.created_datetime.strftime('%Y-%m-%d %H:%M') if app.created_datetime else "",
            app.faculty.name if app.faculty else "",
            app.study_type.name_uz if app.study_type else "",
            "✅" if student and student.is_passed_exam else "❌",
        ]
        for col_num, value in enumerate(row, 1):
            ws.cell(row=idx + 2, column=col_num, value=value)

    # Adjust column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 25

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=registered_students.xlsx"
    wb.save(response)
    return response


# Application Admin
@admin.register(Application)
class ApplicationAdmin(SafeBaseModelAdmin, TranslationAdmin):
    list_display = ['id', 'title', 'start_date', 'end_date', 'is_active', 'programs_list']
    search_fields = ['id', 'title', 'description']
    filter_horizontal = ['programs', 'faculties', 'study_types', 'exams']
    date_hierarchy = 'start_date'
    readonly_fields = SafeBaseModelAdmin.readonly_fields + ['programs_list']

    def programs_list(self, obj):
        return ", ".join([p.name for p in obj.programs.all()])

    programs_list.short_description = "Programs"


# Filter for presence of transcript
class TranscriptFilter(admin.SimpleListFilter):
    title = 'Has transcript'
    parameter_name = 'has_transcript'

    def lookups(self, request, model_admin):
        return (
            ('yes', 'Yes'),
            ('no', 'No'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.exclude(transcript__isnull=True).exclude(transcript='')
        if self.value() == 'no':
            return queryset.filter(models.Q(transcript__isnull=True) | models.Q(transcript=''))
        return queryset


def export_student_applications_as_excel(modeladmin, request, queryset):
    fields = [
        ("ID", "id"),
        ("Student", lambda obj: f"{obj.student.first_name} {obj.student.last_name}" if obj.student else ''),
        ("Student Phone", lambda obj: obj.student.user.phone_number if obj.student and obj.student.user else ''),
        ("Program", lambda obj: obj.program.name if obj.program else ''),
        ("Faculty", lambda obj: obj.faculty.name if obj.faculty else ''),
        ("Study Type", lambda obj: obj.study_type.name if obj.study_type else ''),
        ("Application", lambda obj: obj.application.title if obj.application else ''),
        ("Is Transfer", "is_transfer"),
        ("Transfer Level", "transfer_level"),
        ("Score", "score"),
        ("Exam Date", lambda obj: obj.exam_date.date if obj.exam_date else ''),
        ("Exam Result", "exam_result"),
        ("Is Online Exam", "is_online_exam"),
        ("Transcript File", lambda obj: obj.transcript.url if obj.transcript else ''),
        ("Has Transcript", lambda obj: bool(obj.transcript)),
        ("Is Active", "is_active"),
        ("Language", "lang"),
        ("Created", "created_datetime"),
        ("Modified", "modified_datetime"),
        ("UUID", "uuid"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Applications"
    # Write header
    for col_num, (header, _) in enumerate(fields, 1):
        ws.cell(row=1, column=col_num, value=header)
    # Write data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, (_, attr) in enumerate(fields, 1):
            try:
                value = attr(obj) if callable(attr) else getattr(obj, attr, "")
            except Exception:
                value = ''
            ws.cell(row=row_num, column=col_num, value=str(value))
    for i in range(1, len(fields) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=student_applications_export.xlsx'
    wb.save(response)
    return response


export_student_applications_as_excel.short_description = "Export selected applications to Excel"


@admin.register(StudentApplication)
class StudentApplicationAdmin(SafeBaseModelAdmin):
    list_display = [
        'id', 'student_link', 'application_link', 'program', 'faculty',
        'study_type', 'is_transfer', 'is_active'
    ]
    search_fields = ['id', 'student__pinfl', 'student__first_name', 'student__last_name', 'student__passport_number',
                     'student__user__phone_number', 'program__name', 'faculty__name']
    list_filter = ['is_transfer', 'is_active', TranscriptFilter]
    readonly_fields = ['created_datetime', 'modified_datetime', 'uuid', 'score', 'download_transcript']
    actions = [export_student_applications_as_excel]

    def student_link(self, obj):
        url = f"/admin/user/user/{obj.student.user_id}/change/"
        full_name = f"{obj.student.first_name} {obj.student.last_name}"
        return format_html('<a href="{}">{}</a>', url, full_name)

    student_link.short_description = "Student"

    def application_link(self, obj):
        url = f"/admin/university/application/{obj.application_id}/change/"
        return format_html('<a href="{}">{}</a>', url, obj.application.title)

    application_link.short_description = "Application"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'download/<str:file_type>/<int:student_id>/',
                self.admin_site.admin_view(self.download_file_view),
                name='download-student-application-file',
            )
        ]
        return custom_urls + urls

    def download_transcript(self, obj):
        if obj.transcript:
            url = reverse('admin:download-student-application-file', args=['transcript', obj.id])
            return format_html(f'<a class="button" href="{url}">📄 Transkriptni yuklab olish</a>')
        return "Transkript mavjud emas"
    download_transcript.short_description = "Transkript"

    def download_file_view(self, request, file_type, student_id):
        if not request.user.is_staff:
            raise PermissionDenied("Sizda bu faylni yuklab olishga ruxsat yo'q.")

        try:
            obj = StudentApplication.objects.get(id=student_id)
        except StudentApplication.DoesNotExist:
            raise Http404("Ariza topilmadi.")

        if file_type == "transcript":
            file_field = obj.transcript  # FileField or ImageField
        else:
            raise Http404("Noto‘g‘ri fayl turi.")

        if not file_field or not file_field.name:
            raise Http404("Fayl mavjud emas.")

        file_path = file_field.path

        if not os.path.exists(file_path):
            raise Http404("Fayl topilmadi.")

        file_name = os.path.basename(file_path)
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)


# Exam Admin
@admin.register(Exam)
class ExamAdmin(SafeBaseModelAdmin):
    list_display = ['id', 'program', 'faculty', 'is_online', 'exam_date', 'is_active']
    list_filter = ['is_online', 'is_active']
    search_fields = ['id', 'program__name', 'faculty__name']


# Faculty Admin
@admin.register(Faculty)
class FacultyAdmin(SafeBaseModelAdmin, TranslationAdmin):
    actions = [export_registered_students]
    list_display = [
        'id', 'code', 'name', 'program', 'day_price', 'night_price', 'external_price', 'is_active'
    ]
    search_fields = ['id', 'code', 'name', 'program__name', 'description']
    filter_horizontal = ['subjects']
    list_filter = [
        'is_active', 
        'program',
        ('created_datetime', DateRangeFilter),  # 🔥 Date range filter in admin sidebar
    ]
    readonly_fields = SafeBaseModelAdmin.readonly_fields

    fieldsets = (
        (None, {
            'fields': (
                'code', 'name', 'program',
                'day_price', 'night_price', 'external_price',
                'description', 'subjects',
                'is_active', 'created_datetime', 'modified_datetime', 'uuid'
            )
        }),
    )


# Program Admin
@admin.register(Program)
class ProgramAdmin(SafeBaseModelAdmin, TranslationAdmin):
    list_display = ['id', 'name', 'is_active']
    search_fields = ['id', 'name']
    list_filter = ['is_active']


# StudyType Admin
@admin.register(StudyType)
class StudyTypeAdmin(SafeBaseModelAdmin, TranslationAdmin):
    list_display = ['id', 'name', 'is_active']
    search_fields = ['id', 'name']
    list_filter = ['is_active']


# Subject Admin
@admin.register(Subject)
class SubjectAdmin(SafeBaseModelAdmin, TranslationAdmin):
    list_display = ['id', 'name', 'is_active']
    search_fields = ['id', 'name']
    list_filter = ['is_active']


# ExamType Admin
@admin.register(ExamType)
class ExamTypeAdmin(SafeBaseModelAdmin, TranslationAdmin):
    list_display = ['id', 'name', 'is_active']
    search_fields = ['id', 'name']
    list_filter = ['is_active']


# Question Admin
@admin.register(Question)
class QuestionAdmin(SafeBaseModelAdmin):
    list_display = ['id', 'short_question', 'subject', 'lang', 'is_active']
    search_fields = ['id', 'question_text', 'subject__name']
    list_filter = ['is_active', 'subject', 'lang']
    readonly_fields = SafeBaseModelAdmin.readonly_fields + ['short_question']

    def short_question(self, obj):
        return (obj.question_text[:60] + "...") if obj.question_text and len(
            obj.question_text) > 60 else obj.question_text

    short_question.short_description = "Question"


@admin.register(Duration)
class DurationAdmin(admin.ModelAdmin):
    list_display = ['id', 'faculty', 'study_type', 'study_duration']
    search_fields = ['id', 'faculty__name', 'study_type__name', 'study_duration']
    list_filter = ['faculty', 'study_type']



@admin.register(StudyTypeFaculty)
class StudyTypeFacultyAdmin(admin.ModelAdmin):
    list_display = ['id', 'faculty',  'program']
    search_fields = ['id', 'faculty__name', 'study_type__name']
    list_filter = ['faculty', 'study_type']
    filter_horizontal = ['study_type']
